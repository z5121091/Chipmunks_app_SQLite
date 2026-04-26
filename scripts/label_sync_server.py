#!/usr/bin/python
# -*- coding: utf-8 -*-

"""
掌上仓库 - 数据同步服务 V3.2.1
功能：接收手机发送的数据，自动保存为Excel文件
支持：系统托盘、开机自启动、日志轮转
"""

import cherrypy
import json
import datetime
import os
import sys
import socket
import threading
import logging
from logging.handlers import RotatingFileHandler
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side

# ==================== 配置 ====================
# 数据保存根目录
DATA_ROOT = 'D:/数据同步'
# 发货序列号文件路径
SCAN_FILE = f'{DATA_ROOT}/发货序列号.xlsx'
# 入库单文件路径
INBOUND_FILE = f'{DATA_ROOT}/入库单.xlsx'
# 出库单文件路径
OUTBOUND_FILE = f'{DATA_ROOT}/出库单.xlsx'
# 盘点单文件路径
INVENTORY_FILE = f'{DATA_ROOT}/盘点单.xlsx'
# 标签打印数据文件路径
LABELS_FILE = f'{DATA_ROOT}/标签打印.xlsx'
# 物料数据文件路径
MATERIALS_FILE = f'{DATA_ROOT}/物料数据.xlsx'
# 服务端口
SERVER_PORT = 8080
# ==================== 配置结束 ====================

# 获取程序所在目录
if getattr(sys, 'frozen', False):
    # 打包后的exe路径
    APP_DIR = os.path.dirname(sys.executable)
else:
    # 脚本路径
    APP_DIR = os.path.dirname(os.path.abspath(__file__))

# 日志文件配置
LOG_FILE = os.path.join(APP_DIR, '日志记录.log')
# 日志轮转配置：单个文件最大5MB，保留10个备份
MAX_LOG_SIZE = 5 * 1024 * 1024  # 5MB
BACKUP_COUNT = 10

# 初始化日志记录器
def init_logger():
    """初始化带轮转功能的日志记录器"""
    logger = logging.getLogger('SyncService')
    logger.setLevel(logging.INFO)
    
    # 避免重复添加handler
    if not logger.handlers:
        # 创建轮转文件处理器
        rotating_handler = RotatingFileHandler(
            LOG_FILE,
            maxBytes=MAX_LOG_SIZE,
            backupCount=BACKUP_COUNT,
            encoding='utf-8'
        )
        rotating_handler.setLevel(logging.INFO)
        
        # 设置日志格式
        formatter = logging.Formatter('[%(asctime)s] %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
        rotating_handler.setFormatter(formatter)
        
        logger.addHandler(rotating_handler)
    
    return logger

# 初始化全局日志记录器
logger = init_logger()

def log(message):
    """写入日志（同时输出到控制台和文件）"""
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    log_line = f"[{timestamp}] {message}"
    # 写入轮转日志文件
    logger.info(message)
    # 同时输出到控制台
    print(log_line)

def get_ip():
    """获取本机IP地址"""
    try:
        hostname = socket.gethostname()
        ip_address = socket.gethostbyname(hostname)
        return ip_address
    except:
        return "127.0.0.1"

def apply_excel_styles(sheet):
    """为Excel表格添加样式"""
    thin = Side(border_style="thin", color="000000")
    
    max_row = sheet.max_row
    max_col = sheet.max_column
    
    # 标题行样式
    for col in range(1, max_col + 1):
        cell = sheet.cell(row=1, column=col)
        cell.font = Font(name='微软雅黑', color="000000", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    # 数据行样式
    for row in range(2, max_row + 1):
        for col in range(1, max_col + 1):
            cell = sheet.cell(row=row, column=col)
            cell.font = Font(name='微软雅黑', color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    # 自动列宽
    for col in range(1, max_col + 1):
        max_width = 10
        for row in range(1, max_row + 1):
            cell = sheet.cell(row=row, column=col)
            if cell.value:
                width = sum(2 if ord(c) > 127 else 1 for c in str(cell.value))
                max_width = max(max_width, width)
        sheet.column_dimensions[get_column_letter(col)].width = min(max_width + 2, 50)


# ==================== CORS跨域支持 ====================
def enable_cors():
    """启用CORS跨域支持"""
    cherrypy.response.headers['Access-Control-Allow-Origin'] = '*'
    cherrypy.response.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
    cherrypy.response.headers['Access-Control-Allow-Headers'] = 'Content-Type'

cherrypy.tools.cors = cherrypy.Tool('before_handler', enable_cors)


# ==================== 健康检查接口 ====================
class Health:
    """健康检查端点，供APP检测服务是否运行"""
    exposed = True

    def GET(self):
        enable_cors()
        return json.dumps({
            'status': 'ok',
            'service': '掌上仓库同步服务',
            'version': '3.2.1'
        }, ensure_ascii=False)


# ==================== 发货序列号接口 ====================
class Scans:
    exposed = True

    def OPTIONS(self):
        """处理预检请求"""
        enable_cors()
        return ''

    def POST(self, **kwargs):
        enable_cors()
        try:
            content = kwargs.get("content", "unknown content")
            date = datetime.datetime.now().strftime('%Y-%m-%d')
            month = datetime.datetime.now().strftime('%Y-%m')

            if not os.path.exists(SCAN_FILE):
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = month
                sheet.append(["日期", "序列号"])
                self.set_header_styles(sheet)
            else:
                workbook = load_workbook(SCAN_FILE)
                if month not in workbook.sheetnames:
                    sheet = workbook.create_sheet(title=month)
                    sheet.append(["日期", "序列号"])
                    self.set_header_styles(sheet)
                else:
                    sheet = workbook[month]

            row_num = sheet.max_row + 1
            sheet.append([date, content])
            self.set_row_styles(sheet, row_num)
            sheet.column_dimensions[get_column_letter(1)].width = 15
            sheet.column_dimensions[get_column_letter(2)].width = 100
            workbook.save(SCAN_FILE)

            log(f"已保存序列号: {content}")
            return json.dumps({'success': True, 'message': '保存成功', 'date': date, 'content': content}, ensure_ascii=False)
        except Exception as e:
            log(f"错误: {str(e)}")
            return json.dumps({'success': False, 'message': str(e)}, ensure_ascii=False)

    def set_header_styles(self, sheet):
        thin = Side(border_style="thin", color="000000")
        for cell in sheet[1]:
            cell.font = Font(name='微软雅黑', color="000000", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def set_row_styles(self, sheet, row_num):
        thin = Side(border_style="thin", color="000000")
        for col in range(1, 3):
            cell = sheet.cell(row=row_num, column=col)
            cell.font = Font(name='微软雅黑', color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ==================== 入库单接口 ====================
class Inbound:
    exposed = True

    def OPTIONS(self):
        """处理预检请求"""
        enable_cors()
        return ''

    def POST(self, name_suffix=''):
        enable_cors()
        try:
            log("收到入库单数据...")
            
            cl = cherrypy.request.headers.get('Content-Length', 0)
            if cl:
                cl = int(cl)
            file_content = cherrypy.request.body.read(cl) if cl else cherrypy.request.body.read()
            
            if not file_content:
                return json.dumps({'success': False, 'message': '未收到文件内容'}, ensure_ascii=False)
            
            # 生成带日期的文件名
            today = datetime.datetime.now().strftime('%Y-%m-%d')
            if name_suffix:
                filename = f'入库单_{name_suffix}_{today}.xlsx'
            else:
                filename = f'入库单_{today}.xlsx'
            save_path = os.path.join(DATA_ROOT, filename)
            
            temp_file = save_path.replace('.xlsx', '_temp.xlsx')
            with open(temp_file, 'wb') as f:
                f.write(file_content)
            
            # 加载多Sheet工作簿
            workbook = load_workbook(temp_file)
            
            # 为所有Sheet应用样式
            total_rows = 0
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                apply_excel_styles(sheet)
                total_rows += sheet.max_row - 1  # 减去标题行
            
            workbook.save(save_path)
            os.remove(temp_file)
            
            log(f"已保存入库单: {save_path} ({len(workbook.sheetnames)}个Sheet, 共{total_rows}条记录)")
            
            return json.dumps({'success': True, 'message': '保存成功', 'path': save_path, 'count': total_rows}, ensure_ascii=False)
        except Exception as e:
            log(f"错误: {str(e)}")
            return json.dumps({'success': False, 'message': str(e)}, ensure_ascii=False)
    
    def GET(self):
        enable_cors()
        return json.dumps({'success': True, 'message': '服务运行中', 'save_path': INBOUND_FILE}, ensure_ascii=False)


# ==================== 出库单接口 ====================
class Outbound:
    exposed = True

    def OPTIONS(self):
        """处理预检请求"""
        enable_cors()
        return ''

    def POST(self, name_suffix=''):
        enable_cors()
        try:
            log("收到出库单数据...")
            
            cl = cherrypy.request.headers.get('Content-Length', 0)
            if cl:
                cl = int(cl)
            file_content = cherrypy.request.body.read(cl) if cl else cherrypy.request.body.read()
            
            if not file_content:
                return json.dumps({'success': False, 'message': '未收到文件内容'}, ensure_ascii=False)
            
            # 生成带日期的文件名
            today = datetime.datetime.now().strftime('%Y-%m-%d')
            if name_suffix:
                filename = f'出库单_{name_suffix}_{today}.xlsx'
            else:
                filename = f'出库单_{today}.xlsx'
            save_path = os.path.join(DATA_ROOT, filename)
            
            temp_file = save_path.replace('.xlsx', '_temp.xlsx')
            with open(temp_file, 'wb') as f:
                f.write(file_content)
            
            workbook = load_workbook(temp_file)
            sheet = workbook.active
            sheet.title = '出库单'
            apply_excel_styles(sheet)
            workbook.save(save_path)
            os.remove(temp_file)
            
            row_count = sheet.max_row - 1
            log(f"已保存出库单: {save_path} ({row_count}条记录)")
            
            return json.dumps({'success': True, 'message': '保存成功', 'path': save_path, 'count': row_count}, ensure_ascii=False)
        except Exception as e:
            log(f"错误: {str(e)}")
            return json.dumps({'success': False, 'message': str(e)}, ensure_ascii=False)
    
    def GET(self):
        enable_cors()
        return json.dumps({'success': True, 'message': '服务运行中', 'save_path': OUTBOUND_FILE}, ensure_ascii=False)


# ==================== 盘点单接口 ====================
class Inventory:
    exposed = True

    def OPTIONS(self):
        """处理预检请求"""
        enable_cors()
        return ''

    def POST(self, name_suffix=''):
        enable_cors()
        try:
            log("收到盘点单数据...")
            
            cl = cherrypy.request.headers.get('Content-Length', 0)
            if cl:
                cl = int(cl)
            file_content = cherrypy.request.body.read(cl) if cl else cherrypy.request.body.read()
            
            if not file_content:
                return json.dumps({'success': False, 'message': '未收到文件内容'}, ensure_ascii=False)
            
            # 生成文件名：拆包标签固定名称，其他带日期
            today = datetime.datetime.now().strftime('%Y-%m-%d')
            if name_suffix == '拆包标签':
                filename = '拆包标签.xlsx'
            elif name_suffix:
                filename = f'盘点单_{name_suffix}_{today}.xlsx'
            else:
                filename = f'盘点单_{today}.xlsx'
            save_path = os.path.join(DATA_ROOT, filename)
            
            temp_file = save_path.replace('.xlsx', '_temp.xlsx')
            with open(temp_file, 'wb') as f:
                f.write(file_content)
            
            workbook = load_workbook(temp_file)
            sheet = workbook.active
            sheet.title = '盘点单'
            apply_excel_styles(sheet)
            workbook.save(save_path)
            os.remove(temp_file)
            
            row_count = sheet.max_row - 1
            log(f"已保存盘点单: {save_path} ({row_count}条记录)")
            
            return json.dumps({'success': True, 'message': '保存成功', 'path': save_path, 'count': row_count}, ensure_ascii=False)
        except Exception as e:
            log(f"错误: {str(e)}")
            return json.dumps({'success': False, 'message': str(e)}, ensure_ascii=False)
    
    def GET(self):
        enable_cors()
        return json.dumps({'success': True, 'message': '服务运行中', 'save_path': INVENTORY_FILE}, ensure_ascii=False)


# ==================== 标签打印接口 ====================
class Labels:
    exposed = True

    def OPTIONS(self):
        """处理预检请求"""
        enable_cors()
        return ''

    def POST(self, name_suffix=''):
        enable_cors()
        try:
            log("收到标签数据...")
            
            cl = cherrypy.request.headers.get('Content-Length', 0)
            if cl:
                cl = int(cl)
            file_content = cherrypy.request.body.read(cl) if cl else cherrypy.request.body.read()
            
            if not file_content:
                return json.dumps({'success': False, 'message': '未收到文件内容'}, ensure_ascii=False)
            
            # 生成文件名
            filename = '标签打印.xlsx'
            save_path = os.path.join(DATA_ROOT, filename)
            
            temp_file = save_path.replace('.xlsx', '_temp.xlsx')
            with open(temp_file, 'wb') as f:
                f.write(file_content)
            
            workbook = load_workbook(temp_file)
            sheet = workbook.active
            sheet.title = '标签数据'
            apply_excel_styles(sheet)
            workbook.save(save_path)
            os.remove(temp_file)
            
            row_count = sheet.max_row - 1
            log(f"已保存标签数据: {save_path} ({row_count}条记录)")
            
            return json.dumps({'success': True, 'message': '保存成功', 'path': save_path, 'count': row_count}, ensure_ascii=False)
        except Exception as e:
            log(f"错误: {str(e)}")
            return json.dumps({'success': False, 'message': str(e)}, ensure_ascii=False)
    
    def GET(self):
        enable_cors()
        return json.dumps({'success': True, 'message': '服务运行中', 'save_path': LABELS_FILE}, ensure_ascii=False)


# ==================== 物料数据接口 ====================
class Materials:
    exposed = True

    def OPTIONS(self):
        """处理预检请求"""
        enable_cors()
        return ''

    def POST(self, name_suffix=''):
        enable_cors()
        try:
            log("收到物料数据...")
            
            cl = cherrypy.request.headers.get('Content-Length', 0)
            if cl:
                cl = int(cl)
            file_content = cherrypy.request.body.read(cl) if cl else cherrypy.request.body.read()
            
            if not file_content:
                return json.dumps({'success': False, 'message': '未收到文件内容'}, ensure_ascii=False)
            
            # 生成带日期的文件名
            today = datetime.datetime.now().strftime('%Y-%m-%d')
            if name_suffix:
                filename = f'物料数据_{name_suffix}_{today}.xlsx'
            else:
                filename = f'物料数据_{today}.xlsx'
            save_path = os.path.join(DATA_ROOT, filename)
            
            temp_file = save_path.replace('.xlsx', '_temp.xlsx')
            with open(temp_file, 'wb') as f:
                f.write(file_content)
            
            workbook = load_workbook(temp_file)
            sheet = workbook.active
            sheet.title = '物料数据'
            apply_excel_styles(sheet)
            workbook.save(save_path)
            os.remove(temp_file)
            
            row_count = sheet.max_row - 1
            log(f"已保存物料数据: {save_path} ({row_count}条记录)")
            
            return json.dumps({'success': True, 'message': '保存成功', 'path': save_path, 'count': row_count}, ensure_ascii=False)
        except Exception as e:
            log(f"错误: {str(e)}")
            return json.dumps({'success': False, 'message': str(e)}, ensure_ascii=False)
    
    def GET(self):
        enable_cors()
        return json.dumps({'success': True, 'message': '服务运行中', 'save_path': MATERIALS_FILE}, ensure_ascii=False)


# ==================== 托盘图标 ====================
class TrayIcon:
    def __init__(self):
        self.icon = None
        self.running = True
        self.ip_address = get_ip()
        
    def create_icon_image(self):
        """创建托盘图标"""
        try:
            from PIL import Image, ImageDraw, ImageFont
            
            # 创建一个32x32的图标
            size = 64
            img = Image.new('RGBA', (size, size), (0, 0, 0, 0))
            draw = ImageDraw.Draw(img)
            
            # 绘制圆形背景
            draw.ellipse([4, 4, size-4, size-4], fill=(79, 70, 229, 255))
            
            # 绘制同步符号（两个箭头）
            # 上箭头
            draw.polygon([(size//2, 12), (size//2-8, 24), (size//2+8, 24)], fill=(255, 255, 255, 255))
            # 下箭头
            draw.polygon([(size//2, size-12), (size//2-8, size-24), (size//2+8, size-24)], fill=(255, 255, 255, 255))
            
            return img
        except Exception as e:
            log(f"创建图标失败: {e}")
            # 返回一个简单的图标
            from PIL import Image
            return Image.new('RGBA', (64, 64), (79, 70, 229, 255))
    
    def open_folder(self, folder_path):
        """打开文件夹"""
        try:
            os.makedirs(folder_path, exist_ok=True)
            os.startfile(folder_path)
        except Exception as e:
            log(f"打开文件夹失败: {e}")
    
    def open_logs(self):
        """打开日志文件"""
        try:
            if os.path.exists(LOG_FILE):
                os.startfile(LOG_FILE)
            else:
                log("日志文件不存在")
        except Exception as e:
            log(f"打开日志失败: {e}")
    
    def set_autostart(self, enable=True):
        """设置开机自启动"""
        try:
            import winreg
            
            key_path = r"Software\Microsoft\Windows\CurrentVersion\Run"
            app_name = "LabelSyncService"
            exe_path = sys.executable if getattr(sys, 'frozen', False) else __file__
            
            key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, key_path, 0, winreg.KEY_WRITE)
            
            if enable:
                winreg.SetValueEx(key, app_name, 0, winreg.REG_SZ, f'"{exe_path}"')
                log("已启用开机自启动")
            else:
                try:
                    winreg.DeleteValue(key, app_name)
                    log("已禁用开机自启动")
                except:
                    pass
            
            winreg.CloseKey(key)
        except Exception as e:
            log(f"设置自启动失败: {e}")
    
    def check_autostart(self):
        """检查是否已设置开机自启动"""
        try:
            import winreg
            key_path = r"Software\Microsoft\Windows\CurrentVersion\Run"
            key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, key_path, 0, winreg.KEY_READ)
            try:
                winreg.QueryValueEx(key, "LabelSyncService")
                winreg.CloseKey(key)
                return True
            except:
                winreg.CloseKey(key)
                return False
        except:
            return False
    
    def quit_app(self):
        """退出应用"""
        self.running = False
        if self.icon:
            self.icon.stop()
        cherrypy.engine.exit()
    
    def run(self):
        """运行托盘图标"""
        try:
            import pystray
            from PIL import Image
            
            # 创建图标
            icon_image = self.create_icon_image()
            
            # 创建菜单
            is_autostart = self.check_autostart()
            
            menu = pystray.Menu(
                pystray.MenuItem(
                    lambda text: f"✓ 开机自启动" if self.check_autostart() else "○ 开机自启动",
                    lambda: self.set_autostart(not self.check_autostart()),
                ),
                pystray.Menu.SEPARATOR,
                pystray.MenuItem(f"服务器: {self.ip_address}:{SERVER_PORT}", None, enabled=False),
                pystray.MenuItem(f"数据目录: {DATA_ROOT}", None, enabled=False),
                pystray.Menu.SEPARATOR,
                pystray.MenuItem("打开数据目录", lambda: self.open_folder(DATA_ROOT)),
                pystray.MenuItem("查看日志", self.open_logs),
                pystray.Menu.SEPARATOR,
                pystray.MenuItem("退出", self.quit_app),
            )
            
            # 创建托盘图标
            self.icon = pystray.Icon("label_sync", icon_image, "掌上仓库同步服务", menu)
            self.icon.run()
            
        except ImportError:
            log("缺少依赖: pip install pystray Pillow")
            # 无托盘模式运行
            while self.running:
                import time
                time.sleep(1)


# ==================== 启动服务 ====================
def start_server():
    """启动HTTP服务"""
    ip_address = get_ip()
    
    conf = {
        'global': {
            'server.socket_host': '0.0.0.0',
            'server.socket_port': SERVER_PORT,
            'server.thread_pool': 10,
            'engine.autoreload.on': False,
            'log.screen': False,
        }
    }
    
    # 挂载路由，启用CORS
    cherrypy.tree.mount(Health(), '/health', {
        '/': {
            'request.dispatch': cherrypy.dispatch.MethodDispatcher(),
            'tools.cors.on': True,
        }
    })
    cherrypy.tree.mount(Scans(), '/scans', {
        '/': {
            'request.dispatch': cherrypy.dispatch.MethodDispatcher(),
            'tools.cors.on': True,
        }
    })
    cherrypy.tree.mount(Inbound(), '/inbound', {
        '/': {
            'request.dispatch': cherrypy.dispatch.MethodDispatcher(),
            'tools.cors.on': True,
        }
    })
    cherrypy.tree.mount(Outbound(), '/outbound', {
        '/': {
            'request.dispatch': cherrypy.dispatch.MethodDispatcher(),
            'tools.cors.on': True,
        }
    })
    cherrypy.tree.mount(Inventory(), '/inventory', {
        '/': {
            'request.dispatch': cherrypy.dispatch.MethodDispatcher(),
            'tools.cors.on': True,
        }
    })
    cherrypy.tree.mount(Labels(), '/labels', {
        '/': {
            'request.dispatch': cherrypy.dispatch.MethodDispatcher(),
            'tools.cors.on': True,
        }
    })
    cherrypy.tree.mount(Materials(), '/materials', {
        '/': {
            'request.dispatch': cherrypy.dispatch.MethodDispatcher(),
            'tools.cors.on': True,
        }
    })
    cherrypy.config.update(conf)
    
    log("=" * 50)
    log("  掌上仓库 - 数据同步服务 V3.2.1")
    log("=" * 50)
    log(f"  本机IP: {ip_address}")
    log(f"  服务端口: {SERVER_PORT}")
    log(f"  数据目录: {DATA_ROOT}")
    log("=" * 50)
    
    cherrypy.engine.start()


def main():
    """主函数"""
    # 确保数据目录存在
    os.makedirs(DATA_ROOT, exist_ok=True)
    
    # 启动HTTP服务（在后台线程）
    server_thread = threading.Thread(target=start_server, daemon=True)
    server_thread.start()
    
    # 运行托盘图标（主线程）
    tray = TrayIcon()
    tray.run()


if __name__ == '__main__':
    main()
